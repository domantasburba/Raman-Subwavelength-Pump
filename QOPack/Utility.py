import numpy as np
import cProfile, pstats, io
from datetime import datetime
import logging
import openpyxl
import os
import pathlib
import pandas as pd
import time
import matplotlib as mpl
import matplotlib.pyplot as plt

plt.style.use("default")


# Taken from https://www.youtube.com/watch?v=8qEnExGLZfY.
def profile(fnc):
    """A decorator that uses cProfile to profile a function."""

    def inner(*args, **kwargs):

        pr = cProfile.Profile()
        pr.enable()
        retval = fnc(*args, **kwargs)
        pr.disable()
        s = io.StringIO()
        sortby = "cumulative"
        ps = pstats.Stats(pr, stream=s).sort_stats(sortby).reverse_order()
        ps.print_stats()
        print(s.getvalue())
        return retval

    return inner


def fun_time(fun):
    def wrap_fun(*args, **kwargs):
        start_time = time.perf_counter()
        result = fun(*args, **kwargs)
        end_time = time.perf_counter()

        time_diff = end_time - start_time
        print("%s DONE IN" % fun.__name__)
        print("--- %.4f seconds ---" % time_diff)

        return result

    return wrap_fun


# Taken from https://stackoverflow.com/questions/918154/relative-paths-in-python
def get_main_dir():
    return os.path.dirname(os.path.realpath(__import__("__main__").__file__))
    # return pathlib.Path.cwd()


def get_main_path():
    return os.path.realpath(__import__("__main__").__file__)


def get_main_name():
    """WITHOUT file extension .py"""
    return get_main_path().split('/')[-1].split('.')[0]


def save_log(saveLog, logging_level):
    if saveLog:
        datetime_now = datetime.now()
        # path_string = "%s/Logs/%s" % (get_main_dir(), datetime_now.strftime("%Y-%m-%d"))
        path_string = "%s/Logs/%s" % (get_main_dir(), datetime_now.strftime("%Y-%m-%d"))
        pathlib.Path(path_string).mkdir(parents=True, exist_ok=True)
        name_string = "Log!%s.log" % datetime_now.strftime("%Y-%m-%d!%H.%M.%S")

        logging.basicConfig(level=logging_level, format="%(message)s", handlers=
        [logging.FileHandler("%s/%s" % (path_string, name_string)), logging.StreamHandler()])
    else:
        logging.basicConfig(level=logging_level, format="%(message)s", handlers=[logging.StreamHandler()])


def get_fig_ax(fig_ax=None, fig_name=None):
    """Used to create flexible single graph functions, which can easily be used
    to create composite plots containing multiple graphs.\n
    fig_ax is either None or a tuple of matplotlib's figure and axis objects.
    If fig_ax is None, a single graph will be plotted. If fig_ax is given, the
    graph will be added to the composite plot.\n
    fig_name is either None or a string. fig_name only takes effect if fig_ax is
    None, i.e., a single graph is plotted. fig_name sets the figure name."""
    if fig_ax is None:
        if fig_name is None:
            fig = plt.figure()
        else:
            fig = plt.figure(fig_name)
        ax = plt.axes()
    else:
        fig, ax = fig_ax

    return fig, ax


def set_plot_defaults(fig, ax, addGrid=True):
    fig.tight_layout()
    # plt.subplots() returns numpy array of axes
    if isinstance(ax, np.ndarray):
        ax_1D = np.ravel(ax)
        [ax_i.minorticks_on() for ax_i in ax_1D]
        [ax_i.tick_params(axis="both", which="both", direction="in") for ax_i in ax_1D]
        [ax_i.xaxis.set_ticks_position("both") for ax_i in ax_1D]
        [ax_i.yaxis.set_ticks_position("both") for ax_i in ax_1D]
        if addGrid:
            [ax_i.grid() for ax_i in ax_1D]
    else:
        ax.minorticks_on()
        ax.tick_params(axis="both", which="both", direction="in")
        ax.xaxis.set_ticks_position("both")
        ax.yaxis.set_ticks_position("both")
        if addGrid:
            ax.grid()


def save_array_NPY(arr, name):
    path_string = "%s/NPY" % get_main_dir()
    pathlib.Path(path_string).mkdir(parents=True, exist_ok=True)
    np.save('%s/%s.npy' % (path_string, name), arr)


def save_array_CSV(arr, name):
    path_string = "%s/CSV" % get_main_dir()
    pathlib.Path(path_string).mkdir(parents=True, exist_ok=True)
    np.savetxt("%s/%s.csv" % (path_string, name), arr, delimiter=",")


def save_plot(name, file_type="png", save_path=None, dpi=500):
    datetime_now = datetime.now()
    if save_path is None:
        # save_path = "%s/Graphs/%s" % (get_main_dir(), datetime_now.strftime("%Y-%m-%d"))
        save_path = "%s/Graphs/%s/%s" % (get_main_dir(), datetime_now.strftime("%Y-%m-%d"), name)
    pathlib.Path(save_path).mkdir(parents=True, exist_ok=True)

    # name_string = "%s!%s.%s" % (name, datetime_now.strftime("%Y-%m-%d!%H.%M.%S"), file_type)
    # name_string = "%s!%s.%s" % (name, datetime_now.strftime("%Y-%m-%d!%H.%M.%S"), "svg")
    name_string = "%s!%s.%s" % (name, datetime_now.strftime("%Y-%m-%d!%H.%M.%S"), "pdf")

    # plt.savefig("%s/%s" % (save_path, name_string))
    if file_type == "svg":
        plt.savefig("%s/%s" % (save_path, name_string), format="svg", dpi=dpi)
    else:
        plt.savefig("%s/%s" % (save_path, name_string), dpi=dpi)
    # plt.savefig("%s/%s" % (save_path, name_string), format="svg", dpi=dpi)


def save_animation(anim, name, fps=30, file_type="gif"):
    datetime_now = datetime.now()
    # path_string = "%s/Animations/%s" % (get_main_dir(), datetime_now.strftime("%Y-%m-%d"))
    path_string = "%s/Animations/%s/%s" % (get_main_dir(), datetime_now.strftime("%Y-%m-%d"), name)
    pathlib.Path(path_string).mkdir(parents=True, exist_ok=True)
    name_string = "%s!%s.%s" % (name, datetime_now.strftime("%Y-%m-%d!%H.%M.%S"), file_type)
    if file_type == "gif":
        anim.save("%s/%s" % (path_string, name_string), fps=fps, writer="imagemagick")
    elif file_type == "mp4":
        anim.save("%s/%s" % (path_string, name_string), fps=fps, writer="ffmpeg")
    else:
        print("WARNING: Unsupported file type. (from %s)" % save_animation.__name__)


def save_parameter_TXT(optional_string=None, save_path=None):
    """If specified, optional_string is inserted into the file name.\n
    If save_path=None, the file will be saved in the default directory, namely, CURRENT_DIRECTORY_OF_MAIN_SCRIPT/Parameters/CURRENT_DATE.\n
    If save_path is specified, the file will be saved in the save_path (save_path must be an absolute path).\n
    \n
    Assumes:\n
    1) Parameters are all in a single section;\n
    2) Section starts with the comment "# PARAMETERS";\n
    3) Section ends with the comment "# END OF PARAMETERS";\n
    \n
    It is recommended to use this function in the end of a script, so that times between\n
    graphs and parameter file almost match (no calculation time delay, only file writing delay).\n
    """
    datetime_now = datetime.now()
    if save_path is None:
        save_path = r"%s/Parameters/%s" % (get_main_dir(), datetime_now.strftime("%Y-%m-%d"))
    pathlib.Path(save_path).mkdir(parents=True, exist_ok=True)

    file = open(get_main_path(), "r")
    if optional_string is None:
        save = open(r"%s/%s!%s.txt" % (save_path, get_main_name(), datetime_now.strftime("%H.%M.%S")), "w")
    else:
        # Convert optional_string to string, if it isn't a string already.
        if not isinstance(optional_string, str):
            optional_string = str(optional_string)
        save = open(r"%s/%s!%s!%s.txt" % (save_path, get_main_name(), optional_string, datetime_now.strftime("%H.%M.%S")), "w")

    startWrite = False
    for line in file:
        line_stripped = line.strip()
        if line_stripped == r"# PARAMETERS":
            startWrite = True
        if startWrite:
            save.write(line_stripped + '\n')
        if line_stripped == r"# END OF PARAMETERS":
            break

    file.close()
    save.close()


def load_array_NPY(name):
    path_string = "%s/NPY" % get_main_dir()
    arr = np.load('%s/%s.npy' % (path_string, name))
    return arr


def plot_matrix(arr, N_div=1, name="Hamiltonian"):
    """Assumes square matrix. Similar functionality is provided by seaborn's
    heatmap - sns.heatmap"""
    fig = plt.figure(r"Matrix")
    ax = plt.axes()

    abs_arr = np.abs(arr)
    normed_arr = (abs_arr - np.mean(abs_arr)) / np.std(abs_arr)
    plt.imshow(normed_arr, cmap="viridis")
    plt.colorbar(label=r"$(|A| - m_{|A|})$ / $\sigma_{|A|}$", aspect=30)

    N_matrix = arr.shape[0]
    for i in range(1, N_div):
        plt.axvline(i * (N_matrix // N_div), lw=0.5, color="gray")
        plt.axhline(i * (N_matrix // N_div), lw=0.5, color="gray")
    set_plot_defaults(fig, ax, addGrid=False)
    save_plot(name)
    plt.show()


def print_matrix(arr, row_span=(0, 0), col_span=(0, 0)):
    """If row_span = col_span = (0, 0), prints the whole matrix."""
    if row_span == col_span == (0, 0):
        print(pd.DataFrame(arr))
    else:
        # row_range = np.array(range(*row_span))
        # col_range = np.array(range(*col_span))
        # print(pd.DataFrame(arr[row_range[:, np.newaxis], col_range[np.newaxis, :]], index=row_range, columns=col_range))
        # print(pd.DataFrame(arr[row_range[:, np.newaxis], col_range[np.newaxis, :]], index=row_range, columns=col_range))

        # print(pd.DataFrame(arr[row_span[0]:row_span[1], col_span[0]:col_span[1]], index=range(*row_span), columns=range(*col_span)))

        print(pd.DataFrame(arr[slice(*row_span), slice(*col_span)], index=range(*row_span), columns=range(*col_span)))


@fun_time
# Note that openpyxl is not the most performant package. It seems xlsxwriter is faster.
# TODO: Rewrite this in xlsxwriter.
def save_matrix_XLSX(arr, name="Hamiltonian", precision=3, colorCells=True, cmap="viridis"):
    """Write 2D array to XLSX spreadsheet using the openpyxl module."""

    datetime_now = datetime.now()
    path_string = "%s/XLSX/%s" % (get_main_dir(), datetime_now.strftime("%Y-%m-%d"))
    pathlib.Path(path_string).mkdir(parents=True, exist_ok=True)
    name_string = "%s!%s.xlsx" % (name, datetime_now.strftime("%Y-%m-%d!%H.%M.%S"))

    np_arr = np.array(arr)
    real_arr = np.real(np_arr)
    imag_arr = np.imag(np_arr)
    abs_arr = np.abs(np_arr)
    normed_arr = (abs_arr - np.mean(abs_arr)) / np.std(abs_arr)

    colormap = plt.get_cmap(cmap)
    rgba_arr = colormap(normed_arr)

    # Convert RGBA array to hex ARGB.
    # start_time = time.perf_counter()
    hex_color_arr = []
    for i in range(len(rgba_arr)):
        row = []
        for j in range(len(rgba_arr[0])):
            hex_rgba = mpl.colors.to_hex(rgba_arr[i, j], keep_alpha=True)
            hex_argb = hex_rgba[-2:] + hex_rgba[1:-2]
            row.append(hex_argb)
        hex_color_arr.append(row)
    # print(time.perf_counter() - start_time)

    # Save given float array to workbook.
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()

    # start_time = time.perf_counter()
    thin = openpyxl.styles.Side(border_style="thin", color="000000")
    for i in range(np_arr.shape[0]):
        ws_row = []
        for j in range(np_arr.shape[1]):
            cell = openpyxl.cell.WriteOnlyCell(ws)
            cell.border = openpyxl.styles.Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.column_dimensions[openpyxl.utils.get_column_letter(j + 1)].width = 14

            if imag_arr[i, j] == 0.0:
                cell.value = "%.*f" % (precision, real_arr[i, j])
            else:
                if imag_arr[i, j] >= 0:
                    cell.value = "%.*f+%.*fj" % (precision, real_arr[i, j], precision, imag_arr[i, j])
                else:
                    cell.value = "%.*f%.*fj" % (precision, real_arr[i, j], precision, imag_arr[i, j])

            if colorCells:
                cell.fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor=hex_color_arr[i][j])

            ws_row.append(cell)
        ws.append(ws_row)
    # print(time.perf_counter() - start_time)

    wb.save(r"%s/%s" % (path_string, name_string))
